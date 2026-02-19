# excel_io.py
"""Read and write task data from/to Excel files."""

from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook


@dataclass(frozen=True)
class Task:
    task_id: str
    url: str
    instructions: str


def read_tasks(path: str | Path) -> list[Task]:
    """Read tasks from Excel, skipping rows where status is 'success'."""
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    required = ("task_id", "url", "instructions")
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns in {path}: {missing}")

    task_id_col = headers.index("task_id")
    url_col = headers.index("url")
    instructions_col = headers.index("instructions")
    status_col = headers.index("status") if "status" in headers else None

    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[task_id_col] is None:
            continue
        if status_col is not None and row[status_col] == "success":
            continue
        tasks.append(Task(
            task_id=str(row[task_id_col]),
            url=str(row[url_col] or ""),
            instructions=str(row[instructions_col] or ""),
        ))
    return tasks


def update_task_result(
    path: str | Path,
    task_id: str,
    screenshot_link: str,
    status: str,
    error: str,
    explanation: str = "",
    audio_link: str = "",
) -> None:
    """Write task results back to the Excel file."""
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Add result columns if missing
    for col_name in ("screenshot_link", "status", "error", "explanation", "audio_link"):
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)

    task_id_col = headers.index("task_id")
    ss_col = headers.index("screenshot_link") + 1  # openpyxl is 1-indexed
    status_col = headers.index("status") + 1
    error_col = headers.index("error") + 1
    explanation_col = headers.index("explanation") + 1
    audio_col = headers.index("audio_link") + 1

    for row in ws.iter_rows(min_row=2):
        if str(row[task_id_col].value) == task_id:
            ws.cell(row=row[0].row, column=ss_col, value=screenshot_link)
            ws.cell(row=row[0].row, column=status_col, value=status)
            ws.cell(row=row[0].row, column=error_col, value=error)
            ws.cell(row=row[0].row, column=explanation_col, value=explanation)
            ws.cell(row=row[0].row, column=audio_col, value=audio_link or "")
            break
    else:
        raise ValueError(f"Task ID '{task_id}' not found in {path}")

    wb.save(path)
