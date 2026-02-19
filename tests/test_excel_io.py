# tests/test_excel_io.py
import pytest
from openpyxl import Workbook, load_workbook
from excel_io import Task, read_tasks, update_task_result


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample tasks.xlsx for testing."""
    path = tmp_path / "tasks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["task_id", "url", "instructions"])
    ws.append(["T001", "https://example.com", "Click the login button"])
    ws.append(["T002", "https://example.org", "Fill in the search field with 'hello'"])
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_results(tmp_path):
    """Create a tasks.xlsx that already has result columns and one completed task."""
    path = tmp_path / "tasks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["task_id", "url", "instructions", "screenshot_link", "status", "error"])
    ws.append(["T001", "https://example.com", "Click login", "pics/T001.png", "success", ""])
    ws.append(["T002", "https://example.org", "Fill search", "", "", ""])
    wb.save(path)
    return path


def test_read_tasks_returns_all_rows(sample_xlsx):
    tasks = read_tasks(sample_xlsx)
    assert len(tasks) == 2
    assert tasks[0] == Task(task_id="T001", url="https://example.com", instructions="Click the login button")
    assert tasks[1].task_id == "T002"


def test_read_tasks_skips_completed(xlsx_with_results):
    tasks = read_tasks(xlsx_with_results)
    assert len(tasks) == 1
    assert tasks[0].task_id == "T002"


def test_update_task_result_adds_columns(sample_xlsx):
    update_task_result(sample_xlsx, "T001", "pics/T001_123.png", "success", "")
    wb = load_workbook(sample_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "screenshot_link" in headers
    assert "status" in headers
    assert "error" in headers
    # Check T001 row values
    row2 = [cell.value for cell in ws[2]]
    assert row2[headers.index("screenshot_link")] == "pics/T001_123.png"
    assert row2[headers.index("status")] == "success"


def test_update_task_result_writes_error(sample_xlsx):
    update_task_result(sample_xlsx, "T002", "", "failed", "Element not found")
    wb = load_workbook(sample_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    row3 = [cell.value for cell in ws[3]]
    assert row3[headers.index("status")] == "failed"
    assert row3[headers.index("error")] == "Element not found"


def test_update_existing_result_columns(xlsx_with_results):
    update_task_result(xlsx_with_results, "T002", "pics/T002_456.png", "success", "")
    wb = load_workbook(xlsx_with_results)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    row3 = [cell.value for cell in ws[3]]
    assert row3[headers.index("screenshot_link")] == "pics/T002_456.png"
    assert row3[headers.index("status")] == "success"


def test_update_task_result_raises_on_missing_id(sample_xlsx):
    with pytest.raises(ValueError, match="Task ID 'NONEXISTENT' not found"):
        update_task_result(sample_xlsx, "NONEXISTENT", "", "failed", "not found")


def test_update_task_result_writes_audio_link(sample_xlsx):
    update_task_result(sample_xlsx, "T001", "pics/T001_123.png", "success", "", audio_link="audio/T001_123.wav")
    wb = load_workbook(sample_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "audio_link" in headers
    row2 = [cell.value for cell in ws[2]]
    assert row2[headers.index("audio_link")] == "audio/T001_123.wav"


def test_update_task_result_audio_link_empty_by_default(sample_xlsx):
    update_task_result(sample_xlsx, "T001", "pics/T001_123.png", "success", "")
    wb = load_workbook(sample_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "audio_link" in headers
    row2 = [cell.value for cell in ws[2]]
    assert row2[headers.index("audio_link")] is None or row2[headers.index("audio_link")] == ""
